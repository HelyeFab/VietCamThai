#!/usr/bin/env python3
"""Parse the XLSX itinerary into structured JSON (data/itinerary.json).

The "Roadmap Giornaliera" sheet drives everything. Note the 2026 revision:
  - the trip runs Bangkok -> ... -> Hanoi (reversed from the original Hanoi-first plan)
  - dates are 1-31 October 2026 (31 days)
  - a 2-night island stop (Koh Rong Samloem) sits between Siem Reap and Phnom Penh
  - column A ("ORDIN") is now a stage ordinal, NOT a day number, so day numbers
    are assigned sequentially by row order
  - the Riepilogo Tappe / Budget sheets in the workbook may be STALE — the summary
    here is regenerated from the roadmap, not read from that sheet

Usage:
    python3 parse_data.py                      # auto-find newest matching xlsx in ~/Downloads
    python3 parse_data.py --xlsx /path/to.xlsx # explicit workbook
"""

import argparse
import datetime
import json
from pathlib import Path

import openpyxl

OUTPUT_PATH = Path(__file__).parent / "data" / "itinerary.json"
DOWNLOADS = Path.home() / "Downloads"
MONTHS = {1: "Gen", 2: "Feb", 3: "Mar", 4: "Apr", 5: "Mag", 6: "Giu",
          7: "Lug", 8: "Ago", 9: "Set", 10: "Ott", 11: "Nov", 12: "Dic"}
MONTHS_FULL = {"Gen": "Gennaio", "Feb": "Febbraio", "Mar": "Marzo", "Apr": "Aprile",
               "Mag": "Maggio", "Giu": "Giugno", "Lug": "Luglio", "Ago": "Agosto",
               "Set": "Settembre", "Ott": "Ottobre", "Nov": "Novembre", "Dic": "Dicembre"}

# location substring -> stage id (folder). Order matters: more specific first.
LOC_TO_STAGE = [
    ("KOH RONG", "12-koh-rong-samloem"),
    ("SIHANOUK", "12-koh-rong-samloem"),
    ("BANGKOK", "11-bangkok"),
    ("SIEM REAP", "10-siem-reap-angkor"),
    ("PHNOM PENH", "09-phnom-penh"),
    ("CHAU DOC", "08-chau-doc-mekong"),
    ("HO CHI MINH", "07-ho-chi-minh-city"),
    ("HOI AN", "06-hoi-an"),
    ("HU", "05-hue"),
    ("PHONG NHA", "04-phong-nha"),
    ("NINH BINH", "03-ninh-binh"),
    ("LAN HA", "02-baia-lan-ha"),
    ("HANOI", "01-hanoi"),
]

# id -> (display name, country)
STAGE_META = {
    "11-bangkok":          ("Bangkok",            "Thailandia"),
    "10-siem-reap-angkor": ("Siem Reap & Angkor", "Cambogia"),
    "12-koh-rong-samloem": ("Koh Rong Samloem",   "Cambogia"),
    "09-phnom-penh":       ("Phnom Penh",         "Cambogia"),
    "08-chau-doc-mekong":  ("Chau Doc & Mekong",  "Vietnam"),
    "07-ho-chi-minh-city": ("Ho Chi Minh City",   "Vietnam"),
    "06-hoi-an":           ("Hoi An",             "Vietnam"),
    "05-hue":              ("Hué",                "Vietnam"),
    "04-phong-nha":        ("Phong Nha",          "Vietnam"),
    "03-ninh-binh":        ("Ninh Binh",          "Vietnam"),
    "02-baia-lan-ha":      ("Baia di Lan Ha",     "Vietnam"),
    "01-hanoi":            ("Hanoi",              "Vietnam"),
}

# travel order (first-visited -> last)
STAGE_ORDER = [
    "11-bangkok", "10-siem-reap-angkor", "12-koh-rong-samloem", "09-phnom-penh",
    "08-chau-doc-mekong", "07-ho-chi-minh-city", "06-hoi-an", "05-hue",
    "04-phong-nha", "03-ninh-binh", "02-baia-lan-ha", "01-hanoi",
]


def find_xlsx() -> Path:
    matches = sorted(DOWNLOADS.glob("Vietnam_Cambogia_Thailandia_2026*.xlsx"),
                     key=lambda p: p.stat().st_mtime, reverse=True)
    if not matches:
        raise FileNotFoundError(f"No Vietnam_Cambogia_Thailandia_2026*.xlsx found in {DOWNLOADS}")
    return matches[0]


def stage_for(loc: str) -> str:
    u = (loc or "").upper()
    for key, sid in LOC_TO_STAGE:
        if key in u:
            return sid
    raise ValueError(f"no stage mapping for location {loc!r}")


def fmt_date(value) -> str:
    if isinstance(value, datetime.datetime):
        return f"{value.day} {MONTHS.get(value.month, value.month)}"
    return str(value)


def parse_roadmap(ws):
    """Daily roadmap. Header is row 1; data from row 2. Day numbers are sequential."""
    days = []
    n = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[1] is None:  # no date -> skip blank/spacer rows
            continue
        n += 1
        days.append({
            "day_number": n,
            "date": fmt_date(row[1]),
            "location": str(row[2]) if row[2] else "",
            "country_zone": str(row[3]) if row[3] else "",
            "what_to_see": str(row[4]) if row[4] else "",
            "alternative_extra": str(row[5]) if row[5] else "",
            "transport": str(row[6]) if row[6] else "",
            "accommodation_cost": str(row[7]) if row[7] else "",
            "nights": str(row[8]) if row[8] else "",
        })
    return days


def build_stages(days):
    """Group days into destination stages, in travel order."""
    for d in days:
        d["_stage"] = stage_for(d["location"])
    stages = []
    for sid in STAGE_ORDER:
        sdays = [d for d in days if d.get("_stage") == sid]
        if not sdays:
            raise ValueError(f"no days mapped to stage {sid}")
        name, country = STAGE_META[sid]
        stages.append({
            "id": sid,
            "name": name,
            "days": [d["day_number"] for d in sdays],
            "country": country,
            "stage_days": sdays,
        })
    for d in days:
        d.pop("_stage", None)
    return stages


def date_range(sdays):
    a, b = sdays[0]["date"], sdays[-1]["date"]
    return a if a == b else f"{a.split()[0]}–{b}"


def build_summary(stages):
    """Regenerate the per-stage summary from the roadmap (the workbook sheet is stale)."""
    summary = []
    for st in stages:
        sd = st["stage_days"]
        summary.append({
            "dates": date_range(sd),
            "country": st["country"],
            "stage": st["name"],
            "top_attraction": sd[0]["what_to_see"][:60],
            "transport_arrival": sd[0]["transport"][:60],
            "nights": str(len(sd)),
        })
    return summary


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=None, help="path to the workbook")
    args = ap.parse_args()

    xlsx = args.xlsx or find_xlsx()
    wb = openpyxl.load_workbook(str(xlsx))

    roadmap = parse_roadmap(wb["📍 Roadmap Giornaliera"])
    stages = build_stages(roadmap)
    summary = build_summary(stages)

    # Budget and practical notes live in the markdown (guida/00-pianificazione/*.md);
    # the workbook's Budget/Note sheets are stale, so they are intentionally not emitted here.
    first, last = roadmap[0]["date"], roadmap[-1]["date"]
    trip_dates = f"{first.split()[0]} – {last.split()[0]} {MONTHS_FULL[last.split()[1]]} 2026"
    data = {
        "trip": {
            "title": "Vietnam · Cambogia · Thailandia",
            "dates": trip_dates,
            "duration_days": len(roadmap),
            "travelers": 4,
        },
        "stages": stages,
        "summary": summary,
        "all_days": roadmap,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"✓ Parsed {xlsx.name}")
    print(f"✓ {len(roadmap)} days into {len(stages)} stages, "
          f"trip {data['trip']['dates']}")
    for st in stages:
        print(f"    {st['id']:<22} {st['name']:<20} days {st['days']}  ({date_range(st['stage_days'])})")
    print(f"✓ Saved to {OUTPUT_PATH}")
    return data


if __name__ == "__main__":
    main()
